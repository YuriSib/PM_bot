from openpyxl import load_workbook, Workbook
import gspread
import asyncio
from datetime import datetime

import sqlite_comands as sql


async def filling_table():
    wb = Workbook()
    ws = wb.active

    column = 2
    for name in ['Исполнитель', 'Неполных', 'Полных', 'Стоимость неполных', 'Стоимость полных', 'Итого должны']:
        ws.cell(row=2, column=column).value = name
        column += 1

    column = 2
    for name in ['Название', 'Неполных', 'Полных', 'Исполнитель']:
        ws.cell(row=13, column=column).value = name
        column += 1

    user_list = await sql.get_user()
    table_data_list = await sql.get_unverified_table()

    row_table_cnt = 14
    row_work_cnt = 3
    for user in user_list:
        tg_id, user_name = user[0], user[1]

        incomplete_cards_cnt = 0
        complete_cards_cnt = 0
        for table in table_data_list:
            table_name, incomplete_cards, complete_cards = table[1], table[3], table[4]
            if table[5] == tg_id:
                incomplete_cards_cnt += int(incomplete_cards)
                complete_cards_cnt += int(complete_cards)

                ws.cell(row=row_table_cnt, column=2).value = table_name
                ws.cell(row=row_table_cnt, column=3).value = incomplete_cards
                ws.cell(row=row_table_cnt, column=4).value = complete_cards
                ws.cell(row=row_table_cnt, column=5).value = user_name

                row_table_cnt += 1
            else:
                continue

        ws.cell(row=row_work_cnt, column=2).value = user_name
        ws.cell(row=row_work_cnt, column=3).value = incomplete_cards_cnt
        ws.cell(row=row_work_cnt, column=4).value = complete_cards_cnt
        ws.cell(row=row_work_cnt, column=5).value = incomplete_cards_cnt * 6
        ws.cell(row=row_work_cnt, column=6).value = complete_cards_cnt * 9
        ws.cell(row=row_work_cnt, column=7).value = (incomplete_cards_cnt * 6) + (complete_cards_cnt * 9)
        row_work_cnt += 1

    current_date = datetime.now()
    formatted_date = current_date.strftime('%d.%m.%y')
    wb.save(f'Подсчет работ {formatted_date}.xlsx')


if __name__ == "__main__":
    asyncio.run(filling_table())