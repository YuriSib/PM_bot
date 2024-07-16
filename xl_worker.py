from openpyxl import load_workbook, Workbook
import gspread
import asyncio
from datetime import datetime

import sqlite_comands as sql


async def line_breaks(dir_path_):
    print(dir_path_)
    wb = load_workbook(dir_path_)
    ws = wb.active

    col_num = 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value
            if not value or 'писание' in value:
                break
        if col_num == 9:
            return 'Не найдена колонка "Описание"'

    max_row = ws.max_row

    for row in range(2, max_row+1):
        old_value = ws.cell(row=row, column=col_num-1).value
        if not old_value:
            continue
        new_value = old_value.replace('•', ' <br> •').replace('<br>  <br> •', ' <br> •')
        letter_idx = 0
        for letter in new_value:
            if letter == ' ' or letter == '<' or letter == 'b' or letter == 'r' or letter == '>':
                letter_idx += 1
            else:
                break
        new_value = new_value[letter_idx:]
        ws.cell(row=row, column=col_num-1).value = new_value

    wb.save(dir_path_)


async def cards_count(dir_path_):
    wb = load_workbook(dir_path_)
    ws = wb.active

    column_name, column_desc, column_pic = False, False, False
    col_num = 1
    for row in ws.iter_rows(min_row=1, max_row=1, max_col=10):
        for cell in row:
            col_num += 1
            value = cell.value

            if not value:
                break

            if 'Наименование' in value:
                column_name = col_num
            elif 'Описание' in value:
                column_desc = col_num
            elif 'фото' in value:
                column_pic = col_num
    if not column_name or not column_desc or not column_pic:
        return 'Не найдена одна из колонок, "Описание", "Наименование", "Ссылки на фото"'

    max_row = ws.max_row
    min_cost, max_cost, zero_cost = 0, 0, 0
    for row in range(1, max_row+1):
        description = ws.cell(row=row+1, column=column_desc-1).value
        pic_cnt = ws.cell(row=row+1, column=column_pic-1).value

        if not description and not pic_cnt:
            if not ws.cell(row=row+1, column=1).value:
                break
            else:
                zero_cost += 1
                continue
        else:
            if pic_cnt and description:
                cnt_str = description.count('•')
                cnt_photo = len(pic_cnt.split(';'))
                if cnt_str >= 3 and cnt_photo >= 2:
                    max_cost += 1
                else:
                    min_cost += 1
            else:
                print(row)

    return min_cost, max_cost


async def add_user(user_name):
    gc = gspread.service_account(filename='creds.json')
    sheet = gc.open("Work_count").sheet1

    for col in [2, 5, 8, 11, 14, 17, 20, 23, 26, 29, 32, 35]:
        if not sheet.cell(row=1, col=col).value:
            sheet.update_cell(row=1, col=col, value=user_name)
            sheet.update_cell(row=2, col=col, value='min_cost')
            sheet.update_cell(row=2, col=col+1, value='max_cost')
            sheet.update_cell(row=2, col=col+2, value='Таблицы')
            break


async def add_result(min_cost, max_cost, table_name, user):
    gc = gspread.service_account(filename='creds.json')
    sheet = gc.open("Work_count").sheet1

    user_list = sheet.row_values(1)
    user_col = user_list.index(user)+1

    current_date = datetime.now().strftime('%d.%m.%Y')
    date_list = sheet.col_values(1)

    if current_date in date_list:
        cur_row = date_list.index(current_date)
    else:
        rows = sheet.get_all_values()
        cur_row = len(rows)
        sheet.update_cell(row=cur_row+1, col=1, value=current_date)

    last_min = sheet.cell(cur_row + 1, user_col).value
    last_max = sheet.cell(cur_row + 1, user_col + 1).value

    last_table = sheet.cell(cur_row + 1, user_col + 2).value
    if last_table:
        last_table_list = [table + '.xlsx' for table in last_table.split('.xlsx, ')]
        if table_name in last_table_list:
            return False

    if last_table:
        last_table += ', '
    else:
        last_table = ''
    if not last_max:
        last_max, last_min = 0, 0
    sheet.update_cell(row=cur_row + 1, col=user_col, value=int(last_min) + min_cost)
    sheet.update_cell(row=cur_row + 1, col=user_col + 1, value=int(last_max) + max_cost)
    sheet.update_cell(row=cur_row + 1, col=user_col + 2, value=last_table + table_name)
    return True


async def work_cnt(table_path):
    wb = Workbook()
    ws = wb.active

    column = 2
    for name in ['Исполнитель', 'Неполных', 'Полных', 'Стоимость неполных', 'Стоимость полных', 'Итого должны']:
        ws.cell(row=2, column=column).value = name
        column += 1

    column = 2
    for name in ['Название', 'Неполных', 'Полных', 'Исполнитель']:
        ws.cell(row=13, column=column).value = name
        column += 1

    user_list = await sql.get_user()
    table_data_list = await sql.get_unverified_table()

    row_table_cnt = 14
    row_work_cnt = 3
    for user in user_list:
        tg_id, user_name = user[0], user[1]

        incomplete_cards_cnt = 0
        complete_cards_cnt = 0
        for table in table_data_list:
            table_name, incomplete_cards, complete_cards = table[1], table[3], table[4]
            if table[5] == tg_id:
                incomplete_cards_cnt += int(incomplete_cards)
                complete_cards_cnt += int(complete_cards)

                ws.cell(row=row_table_cnt, column=2).value = table_name
                ws.cell(row=row_table_cnt, column=3).value = incomplete_cards
                ws.cell(row=row_table_cnt, column=4).value = complete_cards
                ws.cell(row=row_table_cnt, column=5).value = user_name

                row_table_cnt += 1
            else:
                continue

        ws.cell(row=row_work_cnt, column=2).value = user_name
        ws.cell(row=row_work_cnt, column=3).value = incomplete_cards_cnt
        ws.cell(row=row_work_cnt, column=4).value = complete_cards_cnt
        ws.cell(row=row_work_cnt, column=5).value = incomplete_cards_cnt * 6
        ws.cell(row=row_work_cnt, column=6).value = complete_cards_cnt * 9
        ws.cell(row=row_work_cnt, column=7).value = (incomplete_cards_cnt * 6) + (complete_cards_cnt * 9)
        row_work_cnt += 1

    wb.save(table_path)


if __name__ == "__main__":
    # asyncio.run(add_user('Жорик'))
    # asyncio.run(add_result(25, 50, 'Таблица3(дата).xlsx', 'Жорик'))
    work_count()